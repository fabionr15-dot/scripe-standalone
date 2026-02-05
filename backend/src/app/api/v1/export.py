"""Export API v1 endpoints for Excel, PDF, CSV exports."""

import csv
import io
from datetime import datetime
from enum import Enum
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.auth.middleware import require_auth
from app.auth.models import UserAccount, UserSearch
from app.logging_config import get_logger
from app.storage.db import db
from app.storage.models import Company, Search

logger = get_logger(__name__)

router = APIRouter(prefix="/searches", tags=["export"])


class ExportFormat(str, Enum):
    """Export format options."""
    CSV = "csv"
    EXCEL = "excel"
    PDF = "pdf"


class ExportOptions(BaseModel):
    """Export configuration options."""
    format: ExportFormat = ExportFormat.CSV
    include_scores: bool = True
    include_validation: bool = True
    min_quality: float = 0.0
    columns: list[str] | None = None  # None = all columns


# Default columns for export
DEFAULT_COLUMNS = [
    "company_name",
    "website",
    "phone",
    "alternative_phones",
    "email",
    "address_line",
    "postal_code",
    "city",
    "region",
    "country",
    "category",
]

SCORE_COLUMNS = [
    "quality_score",
    "match_score",
    "confidence_score",
]

VALIDATION_COLUMNS = [
    "phone_validated",
    "email_validated",
    "website_validated",
]


@router.post("/{search_id}/export")
async def export_search(
    search_id: int,
    options: ExportOptions | None = None,
    user: UserAccount = Depends(require_auth),
):
    """Export search results in specified format.

    Supports CSV, Excel, and PDF formats with customizable columns and filters.
    """
    if options is None:
        options = ExportOptions()

    with db.session() as session:
        # Verify ownership
        user_search = session.query(UserSearch).filter(
            UserSearch.search_id == search_id,
            UserSearch.user_id == user.id,
        ).first()
        if not user_search:
            raise HTTPException(status_code=404, detail="Search not found")

        search = session.query(Search).filter(Search.id == search_id).first()
        if not search:
            raise HTTPException(status_code=404, detail="Search not found")

        # Build query with filters
        query = session.query(Company).filter(Company.search_id == search_id)

        if options.min_quality > 0:
            query = query.filter(Company.quality_score >= options.min_quality)

        companies = query.order_by(Company.quality_score.desc()).all()

        if not companies:
            raise HTTPException(status_code=404, detail="No results to export")

        # Determine columns
        columns = options.columns or DEFAULT_COLUMNS.copy()
        if options.include_scores:
            columns.extend(SCORE_COLUMNS)
        if options.include_validation:
            columns.extend(VALIDATION_COLUMNS)

        # Export based on format
        if options.format == ExportFormat.CSV:
            return _export_csv(search, companies, columns)
        elif options.format == ExportFormat.EXCEL:
            return _export_excel(search, companies, columns)
        elif options.format == ExportFormat.PDF:
            return _export_pdf(search, companies, columns)


@router.get("/{search_id}/export/csv")
async def export_search_csv(
    search_id: int,
    min_quality: float = Query(default=0.0, ge=0.0, le=1.0),
    include_scores: bool = True,
    user: UserAccount = Depends(require_auth),
):
    """Quick CSV export endpoint."""
    options = ExportOptions(
        format=ExportFormat.CSV,
        min_quality=min_quality,
        include_scores=include_scores,
    )
    return await export_search(search_id, options, user)


@router.get("/{search_id}/export/excel")
async def export_search_excel(
    search_id: int,
    min_quality: float = Query(default=0.0, ge=0.0, le=1.0),
    include_scores: bool = True,
    user: UserAccount = Depends(require_auth),
):
    """Quick Excel export endpoint."""
    options = ExportOptions(
        format=ExportFormat.EXCEL,
        min_quality=min_quality,
        include_scores=include_scores,
    )
    return await export_search(search_id, options, user)


@router.get("/{search_id}/export/pdf")
async def export_search_pdf(
    search_id: int,
    min_quality: float = Query(default=0.0, ge=0.0, le=1.0),
    user: UserAccount = Depends(require_auth),
):
    """Quick PDF export endpoint."""
    options = ExportOptions(
        format=ExportFormat.PDF,
        min_quality=min_quality,
        include_scores=True,
    )
    return await export_search(search_id, options, user)


# ==================== EXPORT IMPLEMENTATIONS ====================


def _export_csv(search: Search, companies: list[Company], columns: list[str]) -> StreamingResponse:
    """Export as CSV file."""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()

    for company in companies:
        row = _company_to_dict(company, columns)
        writer.writerow(row)

    output.seek(0)

    filename = f"scripe_export_{search.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
        },
    )


def _export_excel(search: Search, companies: list[Company], columns: list[str]) -> StreamingResponse:
    """Export as Excel file with formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel export requires openpyxl. Install with: pip install openpyxl",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Export"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write header row
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=_format_column_name(column))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Write data rows
    for row_idx, company in enumerate(companies, 2):
        row_data = _company_to_dict(company, columns)
        for col_idx, column in enumerate(columns, 1):
            value = row_data.get(column, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

            # Format score columns as percentage
            if column.endswith("_score") and isinstance(value, (int, float)):
                cell.number_format = "0%"
            # Format boolean columns
            elif column.endswith("_validated"):
                cell.value = "Yes" if value else "No"

    # Auto-adjust column widths
    for col_idx, column in enumerate(columns, 1):
        max_length = len(_format_column_name(column))
        for row_idx in range(2, len(companies) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add summary sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary["A1"] = "Search Name"
    ws_summary["B1"] = search.name
    ws_summary["A2"] = "Export Date"
    ws_summary["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_summary["A3"] = "Total Results"
    ws_summary["B3"] = len(companies)
    ws_summary["A4"] = "Average Quality Score"
    avg_quality = sum(c.quality_score or 0 for c in companies) / len(companies) if companies else 0
    ws_summary["B4"] = f"{avg_quality:.1%}"

    # Style summary
    for row in range(1, 5):
        ws_summary.cell(row=row, column=1).font = Font(bold=True)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"scripe_export_{search.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
        },
    )


def _export_pdf(search: Search, companies: list[Company], columns: list[str]) -> StreamingResponse:
    """Export as PDF report."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
        )
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="PDF export requires reportlab. Install with: pip install reportlab",
        )

    output = io.BytesIO()

    # Create document
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1 * cm,
        bottomMargin=1 * cm,
    )

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=18,
        spaceAfter=20,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.gray,
        spaceAfter=10,
    )

    elements = []

    # Title
    elements.append(Paragraph(f"Lead Export: {search.name}", title_style))
    elements.append(
        Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
            f"Total Results: {len(companies)}",
            subtitle_style,
        )
    )
    elements.append(Spacer(1, 20))

    # Limit columns for PDF (too many won't fit)
    pdf_columns = columns[:8]  # Max 8 columns for readability

    # Create table data
    table_data = [[_format_column_name(col) for col in pdf_columns]]

    for company in companies[:100]:  # Limit to 100 rows for PDF
        row_data = _company_to_dict(company, pdf_columns)
        row = []
        for col in pdf_columns:
            value = row_data.get(col, "")
            # Truncate long values
            if isinstance(value, str) and len(value) > 30:
                value = value[:27] + "..."
            # Format scores
            if col.endswith("_score") and isinstance(value, (int, float)):
                value = f"{value:.0%}"
            elif col.endswith("_validated"):
                value = "✓" if value else "✗"
            row.append(value or "-")
        table_data.append(row)

    # Create table
    col_widths = [doc.width / len(pdf_columns)] * len(pdf_columns)
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Table style
    table.setStyle(
        TableStyle(
            [
                # Header
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                # Body
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                # Grid
                ("GRID", (0, 0), (-1, -1), 0.5, colors.gray),
                # Alternating row colors
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
            ]
        )
    )

    elements.append(table)

    # Add note if truncated
    if len(companies) > 100:
        elements.append(Spacer(1, 20))
        elements.append(
            Paragraph(
                f"Note: Showing first 100 of {len(companies)} results. "
                "For full data, use Excel or CSV export.",
                subtitle_style,
            )
        )

    # Build PDF
    doc.build(elements)
    output.seek(0)

    filename = f"scripe_export_{search.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
        },
    )


# ==================== HELPERS ====================


def _company_to_dict(company: Company, columns: list[str]) -> dict[str, Any]:
    """Convert company to dict with only specified columns."""
    # Parse alternative phones from JSON
    import json
    alt_phones = []
    if company.alternative_phones:
        try:
            alt_phones = json.loads(company.alternative_phones)
        except (json.JSONDecodeError, TypeError):
            pass

    all_data = {
        "company_name": company.company_name or "",
        "website": company.website or "",
        "phone": company.phone or "",
        "alternative_phones": ", ".join(alt_phones) if alt_phones else "",
        "email": company.email or "",
        "address_line": company.address_line or "",
        "postal_code": company.postal_code or "",
        "city": company.city or "",
        "region": company.region or "",
        "country": company.country or "",
        "category": company.category or "",
        "company_size": company.company_size or "",
        "employee_count": company.employee_count or "",
        "quality_score": company.quality_score or 0,
        "match_score": company.match_score or 0,
        "confidence_score": company.confidence_score or 0,
        "phone_validated": company.phone_validated or False,
        "email_validated": company.email_validated or False,
        "website_validated": company.website_validated or False,
        "sources_count": company.sources_count or 1,
    }
    return {col: all_data.get(col, "") for col in columns}


def _format_column_name(column: str) -> str:
    """Format column name for display."""
    return column.replace("_", " ").title()
